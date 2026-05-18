from __future__ import annotations

import csv
import json
from io import BytesIO, StringIO
from pathlib import Path
from uuid import UUID

from openai import AsyncOpenAI
from sqlalchemy import delete, select, update
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

from atendia.config import get_settings
from atendia.db.models.knowledge_document import KnowledgeChunk, KnowledgeDocument
from atendia.storage import get_storage_backend
from atendia.tools.embeddings import generate_embeddings_batch

MAX_PDF_PAGES = 100


async def index_document(ctx: dict, document_id: str, **_: object) -> dict:
    settings = get_settings()
    engine = ctx.get("engine") or create_async_engine(settings.database_url)
    session_factory = async_sessionmaker(engine, expire_on_commit=False)
    try:
        async with session_factory() as session:
            doc = (
                await session.execute(
                    select(KnowledgeDocument).where(KnowledgeDocument.id == UUID(document_id))
                )
            ).scalar_one_or_none()
            if doc is None:
                return {"status": "missing"}
            try:
                data = await get_storage_backend().read(str(doc.tenant_id), doc.storage_path)
                chunks = _parse_document_chunks(doc.filename, data)
                embeddings: list[list[float] | None] = [None] * len(chunks)
                if settings.openai_api_key and chunks:
                    client = AsyncOpenAI(api_key=settings.openai_api_key)
                    embedded, _tokens, _cost = await generate_embeddings_batch(
                        client=client, texts=chunks
                    )
                    embeddings = embedded
                await session.execute(
                    delete(KnowledgeChunk).where(KnowledgeChunk.document_id == doc.id)
                )
                for i, chunk in enumerate(chunks):
                    session.add(
                        KnowledgeChunk(
                            document_id=doc.id,
                            tenant_id=doc.tenant_id,
                            chunk_index=i,
                            text=chunk,
                            embedding=embeddings[i],
                        )
                    )
                doc.status = "indexed"
                doc.fragment_count = len(chunks)
                doc.embedded_chunk_count = sum(1 for item in embeddings if item is not None)
                doc.progress_percentage = 100
                doc.error_message = None
                doc.error_count = 0
            except Exception as exc:
                await session.execute(
                    update(KnowledgeDocument)
                    .where(KnowledgeDocument.id == doc.id)
                    .values(status="error", error_message=str(exc)[:1000])
                )
            await session.commit()
    finally:
        if "engine" not in ctx:
            await engine.dispose()
    return {"status": "ok", "document_id": document_id}


def _parse_document_chunks(filename: str, data: bytes) -> list[str]:
    suffix = Path(filename or "").suffix.lower().lstrip(".")
    if suffix == "jsonl":
        text = data.decode("utf-8", errors="ignore")
        chunks: list[str] = []
        for _, line in zip(range(1000), text.splitlines(), strict=False):
            stripped = line.strip()
            if not stripped:
                continue
            try:
                parsed = json.loads(stripped)
            except json.JSONDecodeError:
                chunks.append(stripped)
            else:
                chunks.append(json.dumps(parsed, ensure_ascii=False, separators=(",", ":")))
        return chunks
    if suffix == "json":
        text = data.decode("utf-8", errors="ignore")
        try:
            parsed = json.loads(text)
        except json.JSONDecodeError:
            return _chunk_text(text)
        records = _json_records_for_retrieval(parsed)
        if records:
            return records
        return _chunk_text(json.dumps(parsed, ensure_ascii=False, indent=2))
    return _chunk_text(_parse_document(filename, data))


def _json_records_for_retrieval(parsed: object) -> list[str]:
    if not isinstance(parsed, dict):
        return []
    records = parsed.get("registros_retrieval")
    if not isinstance(records, list):
        return []
    chunks: list[str] = []
    for record in records[:1000]:
        if isinstance(record, dict):
            text = record.get("texto_retrieval")
            if isinstance(text, str) and text.strip():
                chunks.append(text.strip())
            else:
                chunks.append(json.dumps(record, ensure_ascii=False, separators=(",", ":")))
        elif isinstance(record, str) and record.strip():
            chunks.append(record.strip())
    return chunks


def _parse_document(filename: str, data: bytes) -> str:
    suffix = Path(filename or "").suffix.lower().lstrip(".")
    if suffix == "pdf":
        import fitz  # type: ignore[import-not-found]

        with fitz.open(stream=data, filetype="pdf") as pdf:
            if pdf.page_count > MAX_PDF_PAGES:
                raise ValueError(
                    f"PDF has {pdf.page_count} pages; max supported is {MAX_PDF_PAGES}"
                )
            texts = [page.get_text("text") for page in pdf]
        return "\n".join(texts)
    if suffix == "docx":
        from docx import Document  # type: ignore[import-not-found]

        doc = Document(BytesIO(data))
        return "\n".join(p.text for p in doc.paragraphs)
    if suffix in {"xlsx", "xls"}:
        if suffix == "xls":
            return _parse_xls(data)
        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
        lines: list[str] = []
        for ws in wb.worksheets[:10]:
            for row in ws.iter_rows(max_row=1000, values_only=True):
                values = [str(v) for v in row if v is not None]
                if values:
                    lines.append(", ".join(values))
        return "\n".join(lines)
    text = data.decode("utf-8", errors="ignore")
    if suffix in {"csv", "tsv"}:
        dialect = csv.excel_tab if suffix == "tsv" else csv.excel
        rows = csv.reader(StringIO(text))
        if suffix == "tsv":
            rows = csv.reader(StringIO(text), dialect=dialect)
        return "\n".join(
            ", ".join(cell for cell in row) for _, row in zip(range(1000), rows, strict=False)
        )
    if suffix == "jsonl":
        rows: list[str] = []
        for _, line in zip(range(1000), text.splitlines(), strict=False):
            stripped = line.strip()
            if not stripped:
                continue
            try:
                parsed = json.loads(stripped)
            except json.JSONDecodeError:
                rows.append(stripped)
            else:
                rows.append(json.dumps(parsed, ensure_ascii=False, separators=(",", ":")))
        return "\n".join(rows)
    if suffix == "json":
        try:
            parsed = json.loads(text)
        except json.JSONDecodeError:
            return text
        return json.dumps(parsed, ensure_ascii=False, indent=2)
    return text


def _parse_xls(data: bytes) -> str:
    try:
        import xlrd  # type: ignore[import-not-found]
    except ImportError as exc:  # pragma: no cover - depends on optional package install
        raise ValueError("XLS parsing requires the xlrd package; use XLSX or install xlrd") from exc

    workbook = xlrd.open_workbook(file_contents=data)
    lines: list[str] = []
    for sheet in workbook.sheets()[:10]:
        for row_idx in range(min(sheet.nrows, 1000)):
            values = [
                str(value)
                for value in sheet.row_values(row_idx)
                if value not in (None, "")
            ]
            if values:
                lines.append(", ".join(values))
    return "\n".join(lines)


def _chunk_text(text: str, *, chunk_size: int = 2500, overlap: int = 250) -> list[str]:
    cleaned = "\n".join(line.strip() for line in text.splitlines() if line.strip())
    if not cleaned:
        return []
    chunks: list[str] = []
    start = 0
    while start < len(cleaned) and len(chunks) < 500:
        end = min(len(cleaned), start + chunk_size)
        chunks.append(cleaned[start:end])
        if end == len(cleaned):
            break
        start = max(0, end - overlap)
    return chunks
