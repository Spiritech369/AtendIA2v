from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from atendia.knowledge.os.parsers.base import (
    ParsedDocument,
    ParsedSection,
    ParsedTable,
    normalize_text,
    stringify_cell,
)

MAX_XLSX_SHEETS = 25
MAX_XLSX_ROWS_PER_SHEET = 10000


def parse_xlsx(data: bytes, *, filename: str = "") -> ParsedDocument:
    warnings: list[str] = []
    workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
    tables: list[ParsedTable] = []
    sections: list[ParsedSection] = []
    all_lines: list[str] = []
    sheets = workbook.worksheets
    if len(sheets) > MAX_XLSX_SHEETS:
        warnings.append(f"xlsx_sheet_limit:{len(sheets)}>{MAX_XLSX_SHEETS}")
        sheets = sheets[:MAX_XLSX_SHEETS]
    for sheet_index, worksheet in enumerate(sheets, start=1):
        raw_rows: list[list[Any]] = []
        for row in worksheet.iter_rows():
            if not any(cell.value is not None for cell in row):
                continue
            raw_rows.append([cell.value for cell in row])
            if len(raw_rows) > MAX_XLSX_ROWS_PER_SHEET:
                warnings.append(
                    f"xlsx_row_limit:{worksheet.title}:{len(raw_rows)}>{MAX_XLSX_ROWS_PER_SHEET}"
                )
                raw_rows = raw_rows[:MAX_XLSX_ROWS_PER_SHEET]
                break
        if not raw_rows:
            warnings.append(f"empty_sheet:{worksheet.title}")
            continue
        headers, data_rows, has_headers = _headers_and_rows(raw_rows)
        structured_rows: list[dict[str, Any]] = []
        sheet_lines: list[str] = []
        for row_index, row in enumerate(data_rows, start=1):
            record = {
                headers[col_index]: stringify_cell(row[col_index]) if col_index < len(row) else ""
                for col_index in range(len(headers))
            }
            if not any(record.values()):
                continue
            structured_rows.append(record)
            sheet_lines.append(_row_to_text(record, row_index=row_index, sheet=worksheet.title))
        sheet_text = normalize_text("\n".join(sheet_lines))
        if sheet_text:
            sections.append(
                ParsedSection(
                    text=sheet_text,
                    title=f"{filename or 'Workbook'} - {worksheet.title}",
                    metadata={
                        "file_type": "xlsx",
                        "sheet": worksheet.title,
                        "sheet_index": sheet_index,
                        "row_count": len(structured_rows),
                        "headers": headers,
                    },
                )
            )
            all_lines.append(sheet_text)
        tables.append(
            ParsedTable(
                headers=headers,
                rows=structured_rows,
                reference={"sheet": worksheet.title, "sheet_index": sheet_index},
                metadata={"has_headers": has_headers, "file_type": "xlsx"},
            )
        )
    extracted = normalize_text("\n".join(all_lines))
    return ParsedDocument(
        extracted_text=extracted,
        sections=sections,
        tables=tables,
        metadata={
            "filename": filename,
            "file_type": "xlsx",
            "sheet_count": len(sheets),
            "sheets": [sheet.title for sheet in sheets],
        },
        warnings=warnings,
    )


def _headers_and_rows(rows: list[list[Any]]) -> tuple[list[str], list[list[Any]], bool]:
    first = [stringify_cell(cell) for cell in rows[0]]
    has_headers = any(cell and not _looks_numeric(cell) for cell in first)
    if has_headers:
        headers = [cell or f"column_{index + 1}" for index, cell in enumerate(first)]
        return headers, rows[1:], True
    width = max((len(row) for row in rows), default=0)
    return [f"column_{index + 1}" for index in range(width)], rows, False


def _looks_numeric(value: str) -> bool:
    try:
        float(value.replace(",", "").replace("$", ""))
        return True
    except ValueError:
        return False


def _row_to_text(row: dict[str, str], *, row_index: int, sheet: str) -> str:
    parts = [f"{key}: {value}" for key, value in row.items() if value]
    return f"sheet {sheet} row {row_index}: " + "; ".join(parts)
