from __future__ import annotations

import csv
import json
import unicodedata
from collections.abc import Mapping
from copy import deepcopy
from dataclasses import dataclass, field
from datetime import datetime
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from atendia.catalog_runtime import build_catalog_attrs, coerce_payment_plans, normalize_text_list


@dataclass(slots=True)
class CatalogImportRow:
    row_number: int
    sku: str
    name: str
    category: str | None
    active: bool
    status: str
    price_cents: int | None
    stock_status: str
    region: str | None
    branch: str | None
    payment_plans: list[Any]
    attrs: dict[str, Any]
    tags: list[str] = field(default_factory=list)
    expires_at: datetime | None = None
    issues: list[str] = field(default_factory=list)

    @property
    def valid(self) -> bool:
        return not self.issues


@dataclass(slots=True)
class CatalogImportSource:
    detected_format: str
    source_name: str | None
    source_version: str | None
    rows: list[CatalogImportRow]
    issues: list[str] = field(default_factory=list)


@dataclass(slots=True)
class CatalogImportPreviewRow:
    row_number: int
    sku: str
    name: str
    change_type: str
    reasons: list[str] = field(default_factory=list)


@dataclass(slots=True)
class CatalogImportPreview:
    rows: list[CatalogImportPreviewRow]
    create_count: int
    update_count: int
    unchanged_count: int
    invalid_count: int


_FIELD_ALIASES: dict[str, set[str]] = {
    "sku": {"sku", "id", "codigo", "codigo_producto", "clave", "product_id"},
    "name": {"name", "nombre", "modelo", "producto", "product", "title", "modelo_moto"},
    "category": {"category", "categoria"},
    "active": {"active", "activo", "enabled", "habilitado"},
    "status": {"status", "estado"},
    "price_contado_mxn": {
        "price_contado_mxn",
        "precio_contado_mxn",
        "cash_price_mxn",
        "contado",
        "precio_contado",
    },
    "price_lista_mxn": {
        "price_lista_mxn",
        "precio_lista_mxn",
        "list_price_mxn",
        "lista",
        "precio_lista",
    },
    "stock_status": {"stock_status", "stock", "disponibilidad", "inventory_status"},
    "region": {"region", "zona"},
    "branch": {"branch", "sucursal"},
    "aliases": {
        "aliases",
        "alias",
        "alias_normalizados",
        "aliases_modelo_moto",
        "synonyms",
    },
    "payment_plans": {
        "payment_plans",
        "payment_options",
        "planes_credito",
        "planes_credito_normalizados",
        "planes",
        "planes_pago",
    },
    "product_details": {"product_details", "details", "ficha_tecnica", "specs"},
    "promotions": {"promotions", "promociones", "promo"},
    "rules_by_plan": {"rules_by_plan", "reglas_por_plan", "reglas_plan"},
    "vigencia_inicio": {"vigencia_inicio", "inicio_vigencia", "valid_from"},
    "vigencia_fin": {"vigencia_fin", "fin_vigencia", "valid_to", "expires_at"},
    "tags": {"tags", "tags_uso"},
}


def parse_catalog_document(filename: str, data: bytes) -> CatalogImportSource:
    suffix = Path(filename or "").suffix.lower().lstrip(".")
    if suffix == "json":
        return _parse_json_source(data)
    if suffix in {"csv", "tsv"}:
        return _parse_delimited_source(data, delimiter="\t" if suffix == "tsv" else ",")
    if suffix in {"xlsx", "xlsm"}:
        return _parse_xlsx_source(data)
    raise ValueError("Solo se pueden importar catálogos desde JSON, CSV o XLSX.")


def preview_catalog_import(
    source: CatalogImportSource,
    existing_rows: Mapping[str, Any],
) -> CatalogImportPreview:
    rows: list[CatalogImportPreviewRow] = []
    create_count = 0
    update_count = 0
    unchanged_count = 0
    invalid_count = 0
    for row in source.rows:
        if not row.valid:
            invalid_count += 1
            rows.append(
                CatalogImportPreviewRow(
                    row_number=row.row_number,
                    sku=row.sku,
                    name=row.name,
                    change_type="invalid",
                    reasons=list(row.issues),
                )
            )
            continue
        current = existing_rows.get(row.sku)
        if current is None:
            create_count += 1
            rows.append(
                CatalogImportPreviewRow(
                    row_number=row.row_number,
                    sku=row.sku,
                    name=row.name,
                    change_type="create",
                    reasons=[],
                )
            )
            continue
        reasons = _changed_fields(row, current)
        if reasons:
            update_count += 1
            rows.append(
                CatalogImportPreviewRow(
                    row_number=row.row_number,
                    sku=row.sku,
                    name=row.name,
                    change_type="update",
                    reasons=reasons,
                )
            )
        else:
            unchanged_count += 1
            rows.append(
                CatalogImportPreviewRow(
                    row_number=row.row_number,
                    sku=row.sku,
                    name=row.name,
                    change_type="unchanged",
                    reasons=[],
                )
            )
    return CatalogImportPreview(
        rows=rows,
        create_count=create_count,
        update_count=update_count,
        unchanged_count=unchanged_count,
        invalid_count=invalid_count,
    )


def _parse_json_source(data: bytes) -> CatalogImportSource:
    parsed = json.loads(data.decode("utf-8", errors="ignore"))
    meta = parsed if isinstance(parsed, dict) else {}
    source_name, source_version = _source_metadata(meta)
    records = _extract_json_records(parsed)
    rows = [_normalize_record(item, index + 1) for index, item in enumerate(records)]
    issues = _duplicate_sku_issues(rows)
    return CatalogImportSource(
        detected_format="json",
        source_name=source_name,
        source_version=source_version,
        rows=rows,
        issues=issues,
    )


def _parse_delimited_source(data: bytes, *, delimiter: str) -> CatalogImportSource:
    text = data.decode("utf-8", errors="ignore")
    reader = csv.DictReader(StringIO(text), delimiter=delimiter)
    rows = [_normalize_record(row, index + 2) for index, row in enumerate(reader)]
    issues = _duplicate_sku_issues(rows)
    return CatalogImportSource(
        detected_format="csv" if delimiter == "," else "tsv",
        source_name=None,
        source_version=None,
        rows=rows,
        issues=issues,
    )


def _parse_xlsx_source(data: bytes) -> CatalogImportSource:
    workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    iterator = sheet.iter_rows(values_only=True)
    headers = next(iterator, None)
    if not headers:
        return CatalogImportSource(
            detected_format="xlsx",
            source_name=None,
            source_version=None,
            rows=[],
            issues=["El XLSX no contiene encabezados."],
        )
    normalized_headers = [str(value).strip() if value is not None else "" for value in headers]
    rows: list[CatalogImportRow] = []
    for index, values in enumerate(iterator, start=2):
        mapping = {
            normalized_headers[col_index]: value
            for col_index, value in enumerate(values)
            if col_index < len(normalized_headers) and normalized_headers[col_index]
        }
        if any(value not in (None, "") for value in mapping.values()):
            rows.append(_normalize_record(mapping, index))
    issues = _duplicate_sku_issues(rows)
    return CatalogImportSource(
        detected_format="xlsx",
        source_name=None,
        source_version=None,
        rows=rows,
        issues=issues,
    )


def _source_metadata(root: dict[str, Any]) -> tuple[str | None, str | None]:
    for key in ("kb_metadata", "metadata"):
        meta = root.get(key)
        if not isinstance(meta, dict):
            continue
        return (
            _first_text(meta, {"nombre", "nombre_archivo", "name"}),
            _first_text(meta, {"version"}),
        )
    return None, None


def _extract_json_records(parsed: Any) -> list[Mapping[str, Any]]:
    if isinstance(parsed, list):
        return [item for item in parsed if isinstance(item, Mapping)]
    if not isinstance(parsed, dict):
        return []
    preferred = parsed.get("modelos")
    if isinstance(preferred, list):
        return [item for item in preferred if isinstance(item, Mapping)]
    for key in ("items", "catalog", "products", "rows", "data", "records", "registros_retrieval"):
        value = parsed.get(key)
        if isinstance(value, list):
            return [item for item in value if isinstance(item, Mapping)]
    return []


def _normalize_record(raw: Mapping[str, Any], row_number: int) -> CatalogImportRow:
    issues: list[str] = []
    sku = _first_text(raw, _FIELD_ALIASES["sku"]) or ""
    name = _first_text(raw, _FIELD_ALIASES["name"]) or ""
    category = _first_text(raw, _FIELD_ALIASES["category"])
    active = _bool_value(_pick(raw, _FIELD_ALIASES["active"]), default=True)
    status = _first_text(raw, _FIELD_ALIASES["status"]) or "published"
    stock_status = _first_text(raw, _FIELD_ALIASES["stock_status"]) or "unknown"
    region = _first_text(raw, _FIELD_ALIASES["region"])
    branch = _first_text(raw, _FIELD_ALIASES["branch"])
    price_contado = _pick(raw, _FIELD_ALIASES["price_contado_mxn"])
    price_lista = _pick(raw, _FIELD_ALIASES["price_lista_mxn"])
    payment_source = _pick(raw, _FIELD_ALIASES["payment_plans"])
    payment_plans = coerce_payment_plans(payment_source)
    attrs = deepcopy(raw.get("attrs")) if isinstance(raw.get("attrs"), dict) else {}

    aliases = _list_value(_pick(raw, _FIELD_ALIASES["aliases"]))
    if aliases:
        attrs["aliases"] = aliases

    product_details = _dict_value(_pick(raw, _FIELD_ALIASES["product_details"]))
    if product_details:
        attrs["product_details"] = product_details
        attrs.setdefault("ficha_tecnica", deepcopy(product_details))

    promotions = _json_or_scalar(_pick(raw, _FIELD_ALIASES["promotions"]))
    if promotions not in (None, "", [], {}):
        attrs["promotions"] = promotions

    rules_by_plan = _json_or_scalar(_pick(raw, _FIELD_ALIASES["rules_by_plan"]))
    if rules_by_plan not in (None, "", [], {}):
        attrs["rules_by_plan"] = rules_by_plan

    tags = _list_value(_pick(raw, _FIELD_ALIASES["tags"]))
    if tags:
        attrs["tags"] = tags

    for key in ("busqueda_texto", "texto_retrieval", "tipo_registro", "campo_cliente_principal"):
        value = raw.get(key)
        if isinstance(value, str) and value.strip():
            attrs[key] = value.strip()

    vigencia_inicio = _iso_datetime_text(_pick(raw, _FIELD_ALIASES["vigencia_inicio"]))
    vigencia_fin = _iso_datetime_text(_pick(raw, _FIELD_ALIASES["vigencia_fin"]))
    if vigencia_inicio:
        attrs["vigencia_inicio"] = vigencia_inicio
    if vigencia_fin:
        attrs["vigencia_fin"] = vigencia_fin

    if isinstance(payment_source, dict):
        attrs["payment_options"] = deepcopy(payment_source)

    if price_lista is not None:
        attrs["list_price_mxn"] = _decimal_or_original(price_lista)
    if price_contado is not None:
        attrs["cash_price_mxn"] = _decimal_or_original(price_contado)

    if not sku:
        issues.append("Falta SKU/id.")
    if not name:
        issues.append("Falta name/modelo.")

    price_cents = _mxn_to_cents(price_contado)

    return CatalogImportRow(
        row_number=row_number,
        sku=sku,
        name=name,
        category=category,
        active=active,
        status=status,
        price_cents=price_cents,
        stock_status=stock_status,
        region=region,
        branch=branch,
        payment_plans=payment_plans,
        attrs=build_catalog_attrs(
            base_attrs=attrs,
            sku=sku,
            name=name,
            category=category,
            price_cents=price_cents,
            payment_plans=payment_plans,
        ),
        tags=tags,
        expires_at=_parse_datetime(_pick(raw, _FIELD_ALIASES["vigencia_fin"])),
        issues=issues,
    )


def _duplicate_sku_issues(rows: list[CatalogImportRow]) -> list[str]:
    issues: list[str] = []
    seen: dict[str, int] = {}
    for row in rows:
        if not row.sku:
            continue
        first = seen.get(row.sku)
        if first is None:
            seen[row.sku] = row.row_number
            continue
        row.issues.append(f"SKU duplicado; ya apareció en la fila {first}.")
        issues.append(f"SKU duplicado en filas {first} y {row.row_number}: {row.sku}.")
    return issues


def _changed_fields(imported: CatalogImportRow, current: Any) -> list[str]:
    imported_attrs = _attrs_for_diff(imported.attrs)
    current_attrs = _attrs_for_diff(current.attrs or {})
    checks: list[tuple[str, Any, Any]] = [
        ("name", imported.name, current.name),
        ("category", imported.category, current.category),
        ("active", imported.active, current.active),
        ("status", imported.status, current.status),
        ("price_cents", imported.price_cents, current.price_cents),
        ("stock_status", imported.stock_status, current.stock_status),
        ("region", imported.region, current.region),
        ("branch", imported.branch, current.branch),
        ("payment_plans", imported.payment_plans, current.payment_plans or []),
        ("attrs", imported_attrs, current_attrs),
    ]
    changes = [name for name, left, right in checks if _jsonable(left) != _jsonable(right)]
    if imported.expires_at != getattr(current, "expires_at", None):
        changes.append("vigencia_fin")
    return changes


def _attrs_for_diff(attrs: dict[str, Any]) -> dict[str, Any]:
    cleaned = deepcopy(attrs)
    cleaned.pop("catalog_source", None)
    return cleaned


def _jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {
            str(key): _jsonable(val)
            for key, val in sorted(value.items(), key=lambda item: str(item[0]))
        }
    if isinstance(value, list):
        return [_jsonable(item) for item in value]
    if isinstance(value, tuple):
        return [_jsonable(item) for item in value]
    return value


def _pick(raw: Mapping[str, Any], aliases: set[str]) -> Any:
    normalized = {_normalize_key(key): value for key, value in raw.items()}
    for alias in aliases:
        value = normalized.get(_normalize_key(alias))
        if value not in (None, ""):
            return _json_or_scalar(value)
    if "precios_mxn" in raw and isinstance(raw["precios_mxn"], Mapping):
        nested = raw["precios_mxn"]
        if aliases is _FIELD_ALIASES["price_contado_mxn"]:
            return nested.get("contado")
        if aliases is _FIELD_ALIASES["price_lista_mxn"]:
            return nested.get("lista")
    return None


def _first_text(raw: Mapping[str, Any], aliases: set[str]) -> str | None:
    value = _pick(raw, aliases)
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _normalize_key(value: str) -> str:
    decomposed = unicodedata.normalize("NFKD", value)
    ascii_only = "".join(ch for ch in decomposed if not unicodedata.combining(ch))
    return (
        ascii_only.strip().lower().replace(" ", "_").replace("-", "_").replace(".", "_")
    )


def _json_or_scalar(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    trimmed = value.strip()
    if not trimmed:
        return ""
    if trimmed[:1] not in {"{", "["}:
        return trimmed
    try:
        return json.loads(trimmed)
    except json.JSONDecodeError:
        return trimmed


def _list_value(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, str):
        parsed = _json_or_scalar(value)
        if parsed is not value:
            return _list_value(parsed)
    return normalize_text_list(value)


def _dict_value(value: Any) -> dict[str, Any]:
    if isinstance(value, str):
        value = _json_or_scalar(value)
    return deepcopy(value) if isinstance(value, dict) else {}


def _bool_value(value: Any, *, default: bool) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    if isinstance(value, str):
        normalized = value.strip().casefold()
        if normalized in {"true", "1", "yes", "si", "sí", "activo", "active"}:
            return True
        if normalized in {"false", "0", "no", "inactivo", "inactive"}:
            return False
    return default


def _decimal_or_original(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return value
    return int(number) if number.is_integer() else number


def _mxn_to_cents(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return round(number * 100)


def _parse_datetime(value: Any) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        text = text.replace("Z", "+00:00")
        try:
            return datetime.fromisoformat(text)
        except ValueError:
            return None
    return None


def _iso_datetime_text(value: Any) -> str | None:
    parsed = _parse_datetime(value)
    return parsed.isoformat() if parsed is not None else None
