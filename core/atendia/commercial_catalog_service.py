from __future__ import annotations

import csv
from copy import deepcopy
import json
import re
from collections import defaultdict
from collections.abc import Iterable, Mapping, Sequence
from dataclasses import dataclass, field
from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any
from uuid import UUID

from openpyxl import Workbook, load_workbook
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession

from atendia.db.models.commercial_catalog import (
    Catalog,
    CatalogExportJob,
    CatalogImportJob,
    CatalogItem,
    CatalogItemPlan,
    CatalogVersion,
)

CATALOG_STATUSES = {"active", "inactive", "draft", "archived"}
VERSION_STATUSES = {"draft", "published", "archived"}
JOB_IMPORT_STATUSES = {
    "uploaded",
    "previewed",
    "validated",
    "invalid",
    "draft_saved",
    "published",
    "failed",
}
JOB_EXPORT_STATUSES = {"processing", "ready", "failed"}
STOCK_STATUSES = {"available", "unavailable", "unknown", "limited"}
PLAN_FREQUENCIES = {"daily", "weekly", "biweekly", "quincenal", "monthly", "mensual"}
QUOTE_REQUIRED_DEFAULT = True
VERTICAL_KEY_RE = re.compile(r"^[a-z][a-z0-9_]*$")
CURRENCY_RE = re.compile(r"^[A-Z]{3}$")
NAME_RE = re.compile(r"\s+")
IMPORT_KEY_SANITIZE_RE = re.compile(r"[^a-z0-9_]+")

VERTICAL_TEMPLATES: dict[str, dict[str, Any]] = {
    "motorcycles": {
        "vertical": "motorcycles",
        "label": "Motocicletas",
        "requires_pricing": True,
        "fields": [
            "sku",
            "name",
            "brand",
            "model",
            "base_price",
            "list_price",
            "engine_cc",
            "transmission",
            "stock_quantity",
            "plans",
        ],
    },
    "cars": {
        "vertical": "cars",
        "label": "Automóviles",
        "requires_pricing": True,
        "fields": [
            "sku",
            "name",
            "brand",
            "model",
            "year",
            "base_price",
            "list_price",
            "fuel_type",
            "transmission",
            "stock_quantity",
            "plans",
        ],
    },
    "credits": {
        "vertical": "credits",
        "label": "Financing",
        "requires_pricing": True,
        "fields": [
            "sku",
            "name",
            "product_type",
            "base_price",
            "list_price",
            "min_income",
            "approval_window_days",
            "plans",
        ],
    },
    "real_estate": {
        "vertical": "real_estate",
        "label": "Bienes raíces",
        "requires_pricing": True,
        "fields": [
            "sku",
            "name",
            "operation_type",
            "property_type",
            "location",
            "bedrooms",
            "bathrooms",
            "square_meters",
            "base_price",
        ],
    },
    "clinics": {
        "vertical": "clinics",
        "label": "Clínicas",
        "requires_pricing": True,
        "fields": [
            "sku",
            "name",
            "specialty",
            "duration_minutes",
            "base_price",
            "requires_appointment",
            "requires_evaluation",
        ],
    },
    "restaurants": {
        "vertical": "restaurants",
        "label": "Restaurantes",
        "requires_pricing": True,
        "fields": [
            "sku",
            "name",
            "service_type",
            "base_price",
            "list_price",
            "serves_people",
            "availability_window",
        ],
    },
    "services": {
        "vertical": "services",
        "label": "Servicios",
        "requires_pricing": True,
        "fields": [
            "sku",
            "name",
            "service_type",
            "base_price",
            "duration_minutes",
            "delivery_mode",
        ],
    },
    "physical_products": {
        "vertical": "physical_products",
        "label": "Productos físicos",
        "requires_pricing": True,
        "fields": [
            "sku",
            "name",
            "brand",
            "model",
            "base_price",
            "list_price",
            "stock_quantity",
            "branch_id",
        ],
    },
    "commercial_packages": {
        "vertical": "commercial_packages",
        "label": "Paquetes comerciales",
        "requires_pricing": True,
        "fields": [
            "sku",
            "name",
            "base_price",
            "list_price",
            "included_services",
            "plans",
        ],
    },
}

IMPORT_FIELD_ALIASES: dict[str, set[str]] = {
    "sku": {"sku", "id", "codigo", "codigo_producto", "clave"},
    "name": {"name", "nombre", "modelo", "producto", "product"},
    "type": {"type", "tipo"},
    "category": {"category", "categoria"},
    "base_price": {
        "base_price",
        "cash_price",
        "cash price",
        "price",
        "precio_contado_mxn",
        "cash_price_mxn",
        "precio_contado",
        "contado",
    },
    "list_price": {"list_price", "list price", "precio_lista_mxn", "list_price_mxn", "precio_lista", "lista"},
    "stock_status": {"stock_status", "disponibilidad", "stock"},
    "stock_quantity": {"stock_quantity", "stock_qty", "cantidad", "inventario"},
    "branch_id": {"branch_id", "sucursal", "branch"},
    "status": {"status", "estado"},
    "tags_json": {"tags", "etiquetas", "tags_uso"},
}

PLAN_FIELD_ALIASES: dict[str, set[str]] = {
    "plan_name": {"plan_name", "plan", "nombre_plan"},
    "plan_code": {"plan_code", "codigo_plan"},
    "plan_type": {"plan_type", "tipo_plan"},
    "down_payment_amount": {"down_payment_amount", "advance_payment", "initial_payment"},
    "down_payment_percent": {"down_payment_percent"},
    "installment_amount": {"installment_amount", "installment", "periodic_payment"},
    "installment_frequency": {"installment_frequency", "frequency", "cadence"},
    "installment_count": {"installment_count", "payments", "term_count"},
    "term_months": {"term_months", "months"},
}

TOP_LEVEL_ITEM_FIELDS = {
    "sku",
    "name",
    "type",
    "category",
    "base_price",
    "list_price",
    "stock_status",
    "stock_quantity",
    "branch_id",
    "status",
    "tags_json",
}


@dataclass(slots=True)
class ImportedPlan:
    plan_name: str
    plan_code: str
    plan_type: str | None = None
    down_payment_amount: Decimal | None = None
    down_payment_percent: Decimal | None = None
    installment_amount: Decimal | None = None
    installment_frequency: str | None = None
    installment_count: int | None = None
    term_months: int | None = None
    eligibility_rules_json: dict[str, Any] = field(default_factory=dict)
    status: str = "active"


@dataclass(slots=True)
class ImportedItem:
    sku: str
    name: str
    type: str | None = None
    category: str | None = None
    base_price: Decimal | None = None
    list_price: Decimal | None = None
    stock_status: str = "unknown"
    stock_quantity: int | None = None
    branch_id: str | None = None
    status: str = "active"
    attributes_json: dict[str, Any] = field(default_factory=dict)
    ai_rules_json: dict[str, Any] = field(default_factory=dict)
    tags_json: list[str] = field(default_factory=list)
    plans: list[ImportedPlan] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)

    @property
    def valid(self) -> bool:
        return not self.errors


@dataclass(slots=True)
class ImportPreview:
    items: list[ImportedItem]
    rows_total: int
    rows_valid: int
    rows_error: int
    errors: list[str]


@dataclass(slots=True)
class RuntimeCatalogMatch:
    catalog_id: UUID
    catalog_name: str
    vertical: str
    item: dict[str, Any]
    score: int


def _runtime_match_identity(match: RuntimeCatalogMatch) -> tuple[str, ...]:
    item = match.item if isinstance(match.item, Mapping) else {}
    sku = str(item.get("sku") or "").strip().casefold()
    name = str(item.get("name") or "").strip().casefold()
    category = str(item.get("category") or "").strip().casefold()
    base_price = str(item.get("base_price") or "").strip()
    list_price = str(item.get("list_price") or "").strip()
    if sku or name:
        return ("shape", sku, name, category, base_price, list_price)
    item_id = _none_text(item.get("id"))
    if item_id:
        return ("id", item_id.casefold())
    return ("opaque", repr(sorted(item.items())) if isinstance(item, dict) else repr(item))


def _dedupe_runtime_matches(matches: Sequence[RuntimeCatalogMatch]) -> list[RuntimeCatalogMatch]:
    """Collapse identical published items exposed through multiple active catalogs."""
    deduped: list[RuntimeCatalogMatch] = []
    seen: set[tuple[str, ...]] = set()
    for match in matches:
        identity = _runtime_match_identity(match)
        if identity in seen:
            continue
        seen.add(identity)
        deduped.append(match)
    return deduped


def normalize_vertical(value: str) -> str:
    vertical = _normalize_slug(value)
    if not vertical:
        raise ValueError("vertical is required")
    return vertical


def validate_currency(value: str) -> str:
    currency = (value or "").strip().upper()
    if not CURRENCY_RE.fullmatch(currency):
        raise ValueError("currency must be a valid ISO-style 3-letter code")
    return currency


def validate_status(value: str, *, allowed: set[str], field_name: str) -> str:
    normalized = _normalize_slug(value)
    if normalized not in allowed:
        raise ValueError(f"{field_name} must be one of: {', '.join(sorted(allowed))}")
    return normalized


def validate_json_dict(value: Any, *, field_name: str) -> dict[str, Any]:
    if value is None:
        return {}
    if isinstance(value, str):
        try:
            value = json.loads(value)
        except json.JSONDecodeError as exc:
            raise ValueError(f"{field_name} must be valid JSON") from exc
    if not isinstance(value, dict):
        raise ValueError(f"{field_name} must be a JSON object")
    return dict(value)


def validate_json_list(value: Any, *, field_name: str) -> list[Any]:
    if value is None:
        return []
    if isinstance(value, str):
        try:
            value = json.loads(value)
        except json.JSONDecodeError as exc:
            raise ValueError(f"{field_name} must be valid JSON") from exc
    if not isinstance(value, list):
        raise ValueError(f"{field_name} must be a JSON array")
    return list(value)


def validate_price(value: Any, *, field_name: str) -> Decimal | None:
    if value in (None, ""):
        return None
    amount = _decimal_value(value)
    if amount is None:
        raise ValueError(f"{field_name} must be numeric")
    if amount < 0:
        raise ValueError(f"{field_name} must be greater than or equal to 0")
    return amount


def validate_positive_int(value: Any, *, field_name: str, allow_none: bool = True) -> int | None:
    if value in (None, "") and allow_none:
        return None
    number = _int_value(value)
    if number is None or number <= 0:
        raise ValueError(f"{field_name} must be a positive integer")
    return number


def vertical_template(vertical: str) -> dict[str, Any]:
    return VERTICAL_TEMPLATES.get(vertical, {"vertical": vertical, "label": vertical, "requires_pricing": True, "fields": []})


async def next_catalog_version_number(session: AsyncSession, catalog_id: UUID) -> int:
    rows = (
        await session.execute(
            select(CatalogVersion.version_number).where(CatalogVersion.catalog_id == catalog_id)
        )
    ).scalars().all()
    return (max(rows) if rows else 0) + 1


async def load_catalog_authoring_state(
    session: AsyncSession,
    *,
    tenant_id: UUID,
    catalog_id: UUID,
) -> tuple[Catalog, list[CatalogItem], list[CatalogItemPlan]]:
    catalog = (
        await session.execute(
            select(Catalog).where(Catalog.id == catalog_id, Catalog.tenant_id == tenant_id)
        )
    ).scalar_one_or_none()
    if catalog is None:
        raise ValueError("catalog not found")
    items = (
        await session.execute(
            select(CatalogItem)
            .where(CatalogItem.catalog_id == catalog_id, CatalogItem.tenant_id == tenant_id)
            .order_by(CatalogItem.name.asc())
        )
    ).scalars().all()
    item_ids = [item.id for item in items]
    plans: list[CatalogItemPlan] = []
    if item_ids:
        plans = (
            await session.execute(
                select(CatalogItemPlan)
                .where(CatalogItemPlan.catalog_item_id.in_(item_ids), CatalogItemPlan.tenant_id == tenant_id)
                .order_by(CatalogItemPlan.plan_name.asc())
            )
        ).scalars().all()
    return catalog, items, plans


def snapshot_catalog(
    *,
    catalog: Catalog,
    items: Sequence[CatalogItem],
    plans: Sequence[CatalogItemPlan],
) -> dict[str, Any]:
    plans_by_item: dict[UUID, list[dict[str, Any]]] = defaultdict(list)
    for plan in plans:
        plans_by_item[plan.catalog_item_id].append(
            {
                "plan_name": plan.plan_name,
                "plan_code": plan.plan_code,
                "plan_type": plan.plan_type,
                "down_payment_amount": _decimal_text(plan.down_payment_amount),
                "down_payment_percent": _decimal_text(plan.down_payment_percent),
                "installment_amount": _decimal_text(plan.installment_amount),
                "installment_frequency": plan.installment_frequency,
                "installment_count": plan.installment_count,
                "term_months": plan.term_months,
                "eligibility_rules_json": plan.eligibility_rules_json or {},
                "status": plan.status,
            }
        )
    return {
        "catalog": {
            "id": str(catalog.id),
            "tenant_id": str(catalog.tenant_id),
            "name": catalog.name,
            "vertical": catalog.vertical,
            "currency": catalog.currency,
            "status": catalog.status,
            "active_version_id": str(catalog.active_version_id) if catalog.active_version_id else None,
        },
        "items": [
            {
                "id": str(item.id),
                "sku": item.sku,
                "name": item.name,
                "type": item.type,
                "category": item.category,
                "base_price": _decimal_text(item.base_price),
                "list_price": _decimal_text(item.list_price),
                "stock_status": item.stock_status,
                "stock_quantity": item.stock_quantity,
                "branch_id": item.branch_id,
                "status": item.status,
                "attributes_json": item.attributes_json or {},
                "ai_rules_json": item.ai_rules_json or {},
                "tags_json": item.tags_json or [],
                "plans": plans_by_item.get(item.id, []),
            }
            for item in items
        ],
    }


def validate_snapshot_for_publish(snapshot: Mapping[str, Any]) -> list[str]:
    errors: list[str] = []
    catalog = snapshot.get("catalog") if isinstance(snapshot, Mapping) else {}
    items = snapshot.get("items") if isinstance(snapshot, Mapping) else []
    if not isinstance(catalog, Mapping):
        return ["snapshot catalog metadata is invalid"]
    vertical = normalize_vertical(str(catalog.get("vertical") or "generic"))
    template = vertical_template(vertical)
    requires_pricing = bool(template.get("requires_pricing", QUOTE_REQUIRED_DEFAULT))
    if not isinstance(items, list) or not items:
        errors.append("catalog has no items to publish")
        return errors
    for index, raw_item in enumerate(items):
        if not isinstance(raw_item, Mapping):
            errors.append(f"item {index + 1} is invalid")
            continue
        status_value = _normalize_slug(str(raw_item.get("status") or "draft"))
        if status_value != "active":
            continue
        base_price = _decimal_value(raw_item.get("base_price"))
        list_price = _decimal_value(raw_item.get("list_price"))
        plans = raw_item.get("plans") if isinstance(raw_item.get("plans"), list) else []
        active_plans = [
            plan for plan in plans if isinstance(plan, Mapping) and _normalize_slug(str(plan.get("status") or "active")) == "active"
        ]
        has_any_plan_price = any(
            _decimal_value(plan.get("installment_amount")) is not None
            or _decimal_value(plan.get("down_payment_amount")) is not None
            for plan in active_plans
        )
        if requires_pricing and base_price is None and list_price is None and not has_any_plan_price:
            errors.append(
                f"active item {raw_item.get('sku') or raw_item.get('name') or index + 1} has no price or active plan"
            )
    return errors


async def create_catalog_version(
    session: AsyncSession,
    *,
    catalog: Catalog,
    snapshot: dict[str, Any],
    created_by: UUID | None,
    status: str,
    published_at: datetime | None = None,
) -> CatalogVersion:
    version = CatalogVersion(
        tenant_id=catalog.tenant_id,
        catalog_id=catalog.id,
        version_number=await next_catalog_version_number(session, catalog.id),
        status=status,
        snapshot_json=snapshot,
        created_by=created_by,
        published_at=published_at,
    )
    session.add(version)
    await session.flush()
    return version


async def archive_other_versions(session: AsyncSession, *, catalog_id: UUID, keep_id: UUID) -> None:
    versions = (
        await session.execute(select(CatalogVersion).where(CatalogVersion.catalog_id == catalog_id))
    ).scalars().all()
    for version in versions:
        if version.id == keep_id:
            continue
        if version.status == "published":
            version.status = "archived"


async def publish_authoring_catalog(
    session: AsyncSession,
    *,
    tenant_id: UUID,
    catalog_id: UUID,
    actor_user_id: UUID | None,
) -> CatalogVersion:
    catalog, items, plans = await load_catalog_authoring_state(
        session,
        tenant_id=tenant_id,
        catalog_id=catalog_id,
    )
    snapshot = snapshot_catalog(catalog=catalog, items=items, plans=plans)
    errors = validate_snapshot_for_publish(snapshot)
    if errors:
        raise ValueError("; ".join(errors))
    now = datetime.now(UTC)
    version = await create_catalog_version(
        session,
        catalog=catalog,
        snapshot=snapshot,
        created_by=actor_user_id,
        status="published",
        published_at=now,
    )
    await archive_other_versions(session, catalog_id=catalog.id, keep_id=version.id)
    catalog.active_version_id = version.id
    catalog.status = "active"
    catalog.updated_at = now
    return version


async def publish_existing_version(
    session: AsyncSession,
    *,
    tenant_id: UUID,
    catalog_id: UUID,
    version_id: UUID,
) -> CatalogVersion:
    version = (
        await session.execute(
            select(CatalogVersion).where(
                CatalogVersion.id == version_id,
                CatalogVersion.catalog_id == catalog_id,
                CatalogVersion.tenant_id == tenant_id,
            )
        )
    ).scalar_one_or_none()
    if version is None:
        raise ValueError("version not found")
    errors = validate_snapshot_for_publish(version.snapshot_json or {})
    if errors:
        raise ValueError("; ".join(errors))
    catalog = (
        await session.execute(select(Catalog).where(Catalog.id == catalog_id, Catalog.tenant_id == tenant_id))
    ).scalar_one_or_none()
    if catalog is None:
        raise ValueError("catalog not found")
    await restore_authoring_from_snapshot(
        session,
        tenant_id=tenant_id,
        catalog=catalog,
        snapshot=version.snapshot_json or {},
    )
    now = datetime.now(UTC)
    await archive_other_versions(session, catalog_id=catalog_id, keep_id=version.id)
    version.status = "published"
    version.published_at = now
    catalog.active_version_id = version.id
    catalog.status = "active"
    catalog.updated_at = now
    return version


async def rollback_version(
    session: AsyncSession,
    *,
    tenant_id: UUID,
    catalog_id: UUID,
    version_id: UUID,
    actor_user_id: UUID | None,
) -> CatalogVersion:
    version = (
        await session.execute(
            select(CatalogVersion).where(
                CatalogVersion.id == version_id,
                CatalogVersion.catalog_id == catalog_id,
                CatalogVersion.tenant_id == tenant_id,
            )
        )
    ).scalar_one_or_none()
    if version is None:
        raise ValueError("version not found")
    catalog = (
        await session.execute(select(Catalog).where(Catalog.id == catalog_id, Catalog.tenant_id == tenant_id))
    ).scalar_one_or_none()
    if catalog is None:
        raise ValueError("catalog not found")
    errors = validate_snapshot_for_publish(version.snapshot_json or {})
    if errors:
        raise ValueError("; ".join(errors))
    await restore_authoring_from_snapshot(
        session,
        tenant_id=tenant_id,
        catalog=catalog,
        snapshot=version.snapshot_json or {},
    )
    restored_catalog, items, plans = await load_catalog_authoring_state(
        session,
        tenant_id=tenant_id,
        catalog_id=catalog.id,
    )
    snapshot = snapshot_catalog(catalog=restored_catalog, items=items, plans=plans)
    now = datetime.now(UTC)
    new_version = await create_catalog_version(
        session,
        catalog=catalog,
        snapshot=snapshot,
        created_by=actor_user_id,
        status="published",
        published_at=now,
    )
    await archive_other_versions(session, catalog_id=catalog.id, keep_id=new_version.id)
    catalog.active_version_id = new_version.id
    catalog.status = "active"
    catalog.updated_at = now
    return new_version


async def restore_authoring_from_snapshot(
    session: AsyncSession,
    *,
    tenant_id: UUID,
    catalog: Catalog,
    snapshot: Mapping[str, Any],
) -> None:
    existing_item_ids = (
        await session.execute(
            select(CatalogItem.id).where(
                CatalogItem.catalog_id == catalog.id,
                CatalogItem.tenant_id == tenant_id,
            )
        )
    ).scalars().all()
    if existing_item_ids:
        await session.execute(
            delete(CatalogItemPlan).where(
                CatalogItemPlan.catalog_item_id.in_(existing_item_ids),
                CatalogItemPlan.tenant_id == tenant_id,
            )
        )
    await session.execute(
        delete(CatalogItem).where(CatalogItem.catalog_id == catalog.id, CatalogItem.tenant_id == tenant_id)
    )
    items = snapshot.get("items") if isinstance(snapshot.get("items"), list) else []
    for raw_item in items:
        if not isinstance(raw_item, Mapping):
            continue
        item = CatalogItem(
            tenant_id=tenant_id,
            catalog_id=catalog.id,
            sku=str(raw_item.get("sku") or "").strip(),
            name=str(raw_item.get("name") or "").strip(),
            type=_none_text(raw_item.get("type")),
            category=_none_text(raw_item.get("category")),
            base_price=_decimal_value(raw_item.get("base_price")),
            list_price=_decimal_value(raw_item.get("list_price")),
            stock_status=validate_status(
                str(raw_item.get("stock_status") or "unknown"),
                allowed=STOCK_STATUSES,
                field_name="stock_status",
            ),
            stock_quantity=_int_value(raw_item.get("stock_quantity")),
            branch_id=_none_text(raw_item.get("branch_id")),
            status=validate_status(
                str(raw_item.get("status") or "draft"),
                allowed=CATALOG_STATUSES,
                field_name="status",
            ),
            attributes_json=validate_json_dict(raw_item.get("attributes_json"), field_name="attributes_json"),
            ai_rules_json=validate_json_dict(raw_item.get("ai_rules_json"), field_name="ai_rules_json"),
            tags_json=validate_json_list(raw_item.get("tags_json"), field_name="tags_json"),
            updated_at=datetime.now(UTC),
        )
        session.add(item)
        await session.flush()
        raw_plans = raw_item.get("plans") if isinstance(raw_item.get("plans"), list) else []
        for raw_plan in raw_plans:
            if not isinstance(raw_plan, Mapping):
                continue
            session.add(
                CatalogItemPlan(
                    tenant_id=tenant_id,
                    catalog_item_id=item.id,
                    plan_name=str(raw_plan.get("plan_name") or raw_plan.get("plan_code") or "").strip(),
                    plan_code=str(raw_plan.get("plan_code") or raw_plan.get("plan_name") or "").strip(),
                    plan_type=_none_text(raw_plan.get("plan_type")),
                    down_payment_amount=_decimal_value(raw_plan.get("down_payment_amount")),
                    down_payment_percent=_decimal_value(raw_plan.get("down_payment_percent")),
                    installment_amount=_decimal_value(raw_plan.get("installment_amount")),
                    installment_frequency=_none_text(raw_plan.get("installment_frequency")),
                    installment_count=_int_value(raw_plan.get("installment_count")),
                    term_months=_int_value(raw_plan.get("term_months")),
                    eligibility_rules_json=validate_json_dict(
                        raw_plan.get("eligibility_rules_json"),
                        field_name="eligibility_rules_json",
                    ),
                    status=validate_status(
                        str(raw_plan.get("status") or "active"),
                        allowed=CATALOG_STATUSES,
                        field_name="status",
                    ),
                    updated_at=datetime.now(UTC),
                )
            )


async def search_catalog(
    session: AsyncSession,
    *,
    tenant_id: UUID,
    query: str,
    vertical: str | None = None,
    category: str | None = None,
    status: str | None = None,
    limit: int = 10,
) -> list[RuntimeCatalogMatch]:
    stmt = select(Catalog, CatalogVersion).join(
        CatalogVersion,
        Catalog.active_version_id == CatalogVersion.id,
    ).where(
        Catalog.tenant_id == tenant_id,
        Catalog.status == "active",
        CatalogVersion.status == "published",
    )
    if vertical:
        stmt = stmt.where(Catalog.vertical == normalize_vertical(vertical))
    rows = (await session.execute(stmt)).all()
    q = NAME_RE.sub(" ", (query or "").strip().lower())
    q_tokens = [token for token in q.split(" ") if token]
    matches: list[RuntimeCatalogMatch] = []
    for catalog, version in rows:
        snapshot = version.snapshot_json or {}
        items = snapshot.get("items") if isinstance(snapshot, Mapping) else []
        if not isinstance(items, list):
            continue
        for raw_item in items:
            if not isinstance(raw_item, Mapping):
                continue
            item_status = _normalize_slug(str(raw_item.get("status") or "draft"))
            if status and item_status != _normalize_slug(status):
                continue
            if not status and item_status != "active":
                continue
            item_category = _none_text(raw_item.get("category"))
            if category and (item_category or "").casefold() != str(category).casefold():
                continue
            haystack_parts = [
                str(raw_item.get("id") or ""),
                str(raw_item.get("sku") or ""),
                str(raw_item.get("name") or ""),
                str(raw_item.get("type") or ""),
                str(raw_item.get("category") or ""),
            ]
            tags = raw_item.get("tags_json") if isinstance(raw_item.get("tags_json"), list) else []
            haystack_parts.extend(str(tag) for tag in tags)
            haystack = NAME_RE.sub(" ", " ".join(haystack_parts).lower())
            if q_tokens and not all(token in haystack for token in q_tokens):
                continue
            score = 0
            if q and q == str(raw_item.get("id") or "").lower():
                score += 120
            if q and q == str(raw_item.get("sku") or "").lower():
                score += 100
            if q and q == str(raw_item.get("name") or "").lower():
                score += 90
            score += sum(10 for token in q_tokens if token in haystack)
            matches.append(
                RuntimeCatalogMatch(
                    catalog_id=catalog.id,
                    catalog_name=catalog.name,
                    vertical=catalog.vertical,
                    item=dict(raw_item),
                    score=score,
                )
            )
    matches.sort(
        key=lambda item: (-item.score, str(item.item.get("name") or ""), str(item.item.get("sku") or ""))
    )
    return _dedupe_runtime_matches(matches)[:limit]


async def get_catalog_item_for_quote(
    session: AsyncSession,
    *,
    tenant_id: UUID,
    item_id: str | None = None,
    sku: str | None = None,
    name: str | None = None,
    plan_code: str | None = None,
) -> dict[str, Any]:
    matches = await search_catalog(
        session,
        tenant_id=tenant_id,
        query=sku or name or item_id or "",
        limit=20,
    )
    selected = None
    for match in matches:
        raw = match.item
        if item_id and str(raw.get("id") or "") == item_id:
            selected = raw
            break
        if sku and str(raw.get("sku") or "").casefold() == sku.casefold():
            selected = raw
            break
        if name and str(raw.get("name") or "").casefold() == name.casefold():
            selected = raw
            break
    if selected is None and matches:
        selected = matches[0].item
    if selected is None:
        return {
            "status": "not_found",
            "hint": "Pide al cliente el modelo exacto o muestra opciones similares.",
        }
    item_status = _normalize_slug(str(selected.get("status") or "draft"))
    if item_status != "active":
        return {
            "status": "not_found",
            "hint": "Ese producto no está activo para cotización en la versión publicada.",
        }
    ai_rules = validate_json_dict(selected.get("ai_rules_json"), field_name="ai_rules_json")
    if not ai_rules.get("can_quote", True):
        return {
            "status": "not_found",
            "hint": "Ese producto requiere revisión humana antes de cotizar.",
        }
    if ai_rules.get("requires_human_review"):
        return {
            "status": "not_found",
            "hint": "Ese producto requiere revision humana antes de cotizar.",
        }
    plans = selected.get("plans") if isinstance(selected.get("plans"), list) else []
    selected_plan = None
    if plan_code:
        for plan in plans:
            if not isinstance(plan, Mapping):
                continue
            if str(plan.get("plan_code") or "").casefold() == plan_code.casefold():
                selected_plan = plan
                break
    elif ai_rules.get("requires_plan") and plans:
        return {"status": "missing_data", "missing": ["plan_code"]}
    elif plans:
        selected_plan = next(
            (
                plan
                for plan in plans
                if isinstance(plan, Mapping) and _normalize_slug(str(plan.get("status") or "active")) == "active"
            ),
            None,
        )
    if ai_rules.get("requires_plan") and selected_plan is None:
        return {"status": "missing_data", "missing": ["plan_code"]}
    return {
        "status": "ok",
        "item": {
            "sku": selected.get("sku"),
            "name": selected.get("name"),
            "category": selected.get("category"),
            "base_price": _decimal_number(selected.get("base_price")),
            "list_price": _decimal_number(selected.get("list_price")),
            "stock_status": selected.get("stock_status"),
            "attributes": validate_json_dict(selected.get("attributes_json"), field_name="attributes_json"),
        },
        "plan": (
            {
                "plan_code": selected_plan.get("plan_code"),
                "down_payment_amount": _decimal_number(selected_plan.get("down_payment_amount")),
                "installment_amount": _decimal_number(selected_plan.get("installment_amount")),
                "installment_frequency": selected_plan.get("installment_frequency"),
                "installment_count": selected_plan.get("installment_count"),
                "term_months": selected_plan.get("term_months"),
                "eligibility_rules_json": validate_json_dict(
                    selected_plan.get("eligibility_rules_json"),
                    field_name="eligibility_rules_json",
                ),
            }
            if isinstance(selected_plan, Mapping)
            else None
        ),
        "ai_rules": ai_rules,
        "source": {
            "quote_source": "atendia_knowledge_base",
            "catalog_source": "atendia_catalog_published",
        },
    }


async def has_published_catalogs(session: AsyncSession, *, tenant_id: UUID) -> bool:
    row = (
        await session.execute(
            select(Catalog.id)
            .join(CatalogVersion, Catalog.active_version_id == CatalogVersion.id)
            .where(
                Catalog.tenant_id == tenant_id,
                Catalog.status == "active",
                CatalogVersion.status == "published",
            )
            .limit(1)
        )
    ).scalar_one_or_none()
    return row is not None


def parse_import_file(filename: str, data: bytes) -> dict[str, Any]:
    suffix = Path(filename or "").suffix.lower().lstrip(".")
    if suffix == "json":
        parsed = json.loads(data.decode("utf-8", errors="ignore"))
        rows = _extract_json_rows(parsed)
        columns = _collect_columns(rows)
        return {
            "file_type": "json",
            "rows": rows,
            "columns": columns,
            "meta": parsed if isinstance(parsed, Mapping) else {},
        }
    if suffix in {"csv", "tsv"}:
        text = data.decode("utf-8", errors="ignore")
        reader = csv.DictReader(StringIO(text), delimiter="\t" if suffix == "tsv" else ",")
        rows = [dict(row) for row in reader]
        return {"file_type": suffix, "rows": rows, "columns": reader.fieldnames or [], "meta": {}}
    if suffix in {"xlsx", "xlsm"}:
        workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
        sheet = workbook.worksheets[0]
        iterator = sheet.iter_rows(values_only=True)
        headers = next(iterator, None)
        if not headers:
            return {"file_type": "xlsx", "rows": [], "columns": [], "meta": {}}
        columns = [str(value).strip() if value is not None else "" for value in headers]
        rows: list[dict[str, Any]] = []
        for values in iterator:
            row = {
                columns[index]: value
                for index, value in enumerate(values)
                if index < len(columns) and columns[index]
            }
            if any(value not in (None, "") for value in row.values()):
                rows.append(row)
        return {"file_type": "xlsx", "rows": rows, "columns": columns, "meta": {}}
    raise ValueError("Only CSV, XLSX and JSON imports are supported")


def detect_column_mapping(columns: Sequence[str]) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for column in columns:
        normalized = _normalize_import_key(column)
        for target, aliases in IMPORT_FIELD_ALIASES.items():
            if normalized in {_normalize_import_key(alias) for alias in aliases}:
                mapping[column] = target
                break
        if column in mapping:
            continue
        plan_target = _detect_plan_column_target(column)
        if plan_target is not None:
            mapping[column] = plan_target
    return mapping


def transform_import_rows(
    *,
    rows: Sequence[Mapping[str, Any]],
    column_mapping: Mapping[str, str],
    default_status: str,
) -> ImportPreview:
    items: list[ImportedItem] = []
    errors: list[str] = []
    seen_skus: set[str] = set()
    for index, raw_row in enumerate(rows, start=1):
        item, row_errors = _transform_import_row(raw_row, column_mapping, default_status=default_status)
        if item.sku and item.sku in seen_skus:
            row_errors.append(f"row {index}: duplicate SKU {item.sku}")
        if item.sku:
            seen_skus.add(item.sku)
        if not item.valid or row_errors:
            item.errors.extend(row_errors)
            errors.extend(row_errors)
        items.append(item)
    rows_valid = sum(1 for item in items if item.valid)
    return ImportPreview(
        items=items,
        rows_total=len(rows),
        rows_valid=rows_valid,
        rows_error=len(items) - rows_valid,
        errors=errors,
    )


def serialize_import_preview(preview: ImportPreview) -> dict[str, Any]:
    return {
        "rows_total": preview.rows_total,
        "rows_valid": preview.rows_valid,
        "rows_error": preview.rows_error,
        "errors": list(preview.errors),
        "items": [
            {
                "sku": item.sku,
                "name": item.name,
                "type": item.type,
                "category": item.category,
                "base_price": _decimal_text(item.base_price),
                "list_price": _decimal_text(item.list_price),
                "stock_status": item.stock_status,
                "stock_quantity": item.stock_quantity,
                "branch_id": item.branch_id,
                "status": item.status,
                "attributes_json": dict(item.attributes_json),
                "ai_rules_json": dict(item.ai_rules_json),
                "tags_json": list(item.tags_json),
                "plans": [
                    {
                        "plan_name": plan.plan_name,
                        "plan_code": plan.plan_code,
                        "plan_type": plan.plan_type,
                        "down_payment_amount": _decimal_text(plan.down_payment_amount),
                        "down_payment_percent": _decimal_text(plan.down_payment_percent),
                        "installment_amount": _decimal_text(plan.installment_amount),
                        "installment_frequency": plan.installment_frequency,
                        "installment_count": plan.installment_count,
                        "term_months": plan.term_months,
                        "eligibility_rules_json": dict(plan.eligibility_rules_json),
                        "status": plan.status,
                    }
                    for plan in item.plans
                ],
                "errors": list(item.errors),
            }
            for item in preview.items
        ],
    }


def imported_item_snapshot(item: ImportedItem) -> dict[str, Any]:
    return {
        "sku": item.sku,
        "name": item.name,
        "type": item.type,
        "category": item.category,
        "base_price": _decimal_text(item.base_price),
        "list_price": _decimal_text(item.list_price),
        "stock_status": item.stock_status,
        "stock_quantity": item.stock_quantity,
        "branch_id": item.branch_id,
        "status": item.status,
        "attributes_json": dict(item.attributes_json),
        "ai_rules_json": dict(item.ai_rules_json),
        "tags_json": sorted(str(tag) for tag in item.tags_json),
        "plans": [
            {
                "plan_name": plan.plan_name,
                "plan_code": plan.plan_code,
                "plan_type": plan.plan_type,
                "down_payment_amount": _decimal_text(plan.down_payment_amount),
                "down_payment_percent": _decimal_text(plan.down_payment_percent),
                "installment_amount": _decimal_text(plan.installment_amount),
                "installment_frequency": plan.installment_frequency,
                "installment_count": plan.installment_count,
                "term_months": plan.term_months,
                "eligibility_rules_json": dict(plan.eligibility_rules_json),
                "status": plan.status,
            }
            for plan in sorted(item.plans, key=lambda value: (value.plan_code, value.plan_name))
        ],
    }


def canonicalize_snapshot_item(raw_item: Mapping[str, Any]) -> dict[str, Any]:
    plans = raw_item.get("plans") if isinstance(raw_item.get("plans"), list) else []
    canonical_plans = []
    for plan in plans:
        if not isinstance(plan, Mapping):
            continue
        canonical_plans.append(
            {
                "plan_name": _none_text(plan.get("plan_name")),
                "plan_code": _none_text(plan.get("plan_code")),
                "plan_type": _none_text(plan.get("plan_type")),
                "down_payment_amount": _decimal_text(_decimal_value(plan.get("down_payment_amount"))),
                "down_payment_percent": _decimal_text(_decimal_value(plan.get("down_payment_percent"))),
                "installment_amount": _decimal_text(_decimal_value(plan.get("installment_amount"))),
                "installment_frequency": _none_text(plan.get("installment_frequency")),
                "installment_count": _int_value(plan.get("installment_count")),
                "term_months": _int_value(plan.get("term_months")),
                "eligibility_rules_json": validate_json_dict(
                    plan.get("eligibility_rules_json"),
                    field_name="eligibility_rules_json",
                ),
                "status": _normalize_slug(str(plan.get("status") or "draft")),
            }
        )
    canonical_plans.sort(key=lambda value: (value["plan_code"] or "", value["plan_name"] or ""))
    return {
        "sku": _none_text(raw_item.get("sku")) or "",
        "name": _none_text(raw_item.get("name")) or "",
        "type": _none_text(raw_item.get("type")),
        "category": _none_text(raw_item.get("category")),
        "base_price": _decimal_text(_decimal_value(raw_item.get("base_price"))),
        "list_price": _decimal_text(_decimal_value(raw_item.get("list_price"))),
        "stock_status": _normalize_slug(str(raw_item.get("stock_status") or "unknown")),
        "stock_quantity": _int_value(raw_item.get("stock_quantity")),
        "branch_id": _none_text(raw_item.get("branch_id")),
        "status": _normalize_slug(str(raw_item.get("status") or "draft")),
        "attributes_json": validate_json_dict(raw_item.get("attributes_json"), field_name="attributes_json"),
        "ai_rules_json": validate_json_dict(raw_item.get("ai_rules_json"), field_name="ai_rules_json"),
        "tags_json": sorted(
            str(tag) for tag in validate_json_list(raw_item.get("tags_json"), field_name="tags_json")
        ),
        "plans": canonical_plans,
    }


async def apply_import_preview(
    session: AsyncSession,
    *,
    tenant_id: UUID,
    catalog: Catalog,
    import_job: CatalogImportJob,
    preview: ImportPreview,
    mode: str,
    actor_user_id: UUID | None,
    publish: bool,
) -> dict[str, Any]:
    if mode not in {"merge", "replace"}:
        raise ValueError("mode must be merge or replace")
    if preview.rows_error:
        raise ValueError("import preview still has validation errors")
    now = datetime.now(UTC)
    default_status = "active" if publish else "draft"
    existing_items = (
        await session.execute(
            select(CatalogItem).where(CatalogItem.catalog_id == catalog.id, CatalogItem.tenant_id == tenant_id)
        )
    ).scalars().all()
    existing_by_sku = {item.sku: item for item in existing_items}
    touched_item_ids: list[UUID] = []
    if mode == "replace":
        existing_ids = [item.id for item in existing_items]
        if existing_ids:
            await session.execute(
                delete(CatalogItemPlan).where(
                    CatalogItemPlan.catalog_item_id.in_(existing_ids),
                    CatalogItemPlan.tenant_id == tenant_id,
                )
            )
        await session.execute(
            delete(CatalogItem).where(CatalogItem.catalog_id == catalog.id, CatalogItem.tenant_id == tenant_id)
        )
        existing_by_sku = {}
    created_count = 0
    updated_count = 0
    for imported in preview.items:
        item_status = imported.status if imported.status else default_status
        current = existing_by_sku.get(imported.sku)
        if current is None:
            current = CatalogItem(
                tenant_id=tenant_id,
                catalog_id=catalog.id,
                sku=imported.sku,
                name=imported.name,
                type=imported.type,
                category=imported.category,
                base_price=imported.base_price,
                list_price=imported.list_price,
                stock_status=imported.stock_status,
                stock_quantity=imported.stock_quantity,
                branch_id=imported.branch_id,
                status=item_status,
                attributes_json=imported.attributes_json,
                ai_rules_json=imported.ai_rules_json,
                tags_json=imported.tags_json,
                updated_at=now,
            )
            session.add(current)
            await session.flush()
            created_count += 1
        else:
            current.name = imported.name
            current.type = imported.type
            current.category = imported.category
            current.base_price = imported.base_price
            current.list_price = imported.list_price
            current.stock_status = imported.stock_status
            current.stock_quantity = imported.stock_quantity
            current.branch_id = imported.branch_id
            current.status = item_status
            current.attributes_json = imported.attributes_json
            current.ai_rules_json = imported.ai_rules_json
            current.tags_json = imported.tags_json
            current.updated_at = now
            updated_count += 1
        touched_item_ids.append(current.id)
        await session.execute(
            delete(CatalogItemPlan).where(
                CatalogItemPlan.catalog_item_id == current.id,
                CatalogItemPlan.tenant_id == tenant_id,
            )
        )
        for plan in imported.plans:
            session.add(
                CatalogItemPlan(
                    tenant_id=tenant_id,
                    catalog_item_id=current.id,
                    plan_name=plan.plan_name,
                    plan_code=plan.plan_code,
                    plan_type=plan.plan_type,
                    down_payment_amount=plan.down_payment_amount,
                    down_payment_percent=plan.down_payment_percent,
                    installment_amount=plan.installment_amount,
                    installment_frequency=plan.installment_frequency,
                    installment_count=plan.installment_count,
                    term_months=plan.term_months,
                    eligibility_rules_json=plan.eligibility_rules_json,
                    status=plan.status if plan.status else item_status,
                    updated_at=now,
                )
            )
    import_job.rows_total = preview.rows_total
    import_job.rows_valid = preview.rows_valid
    import_job.rows_error = preview.rows_error
    import_job.validation_errors_json = preview.errors
    import_job.status = "published" if publish else "draft_saved"
    catalog.updated_at = now
    if publish:
        version = await publish_authoring_catalog(
            session,
            tenant_id=tenant_id,
            catalog_id=catalog.id,
            actor_user_id=actor_user_id,
        )
        return {
            "created_count": created_count,
            "updated_count": updated_count,
            "published_version_id": str(version.id),
            "status": "published",
        }
    draft_catalog, items, plans = await load_catalog_authoring_state(
        session,
        tenant_id=tenant_id,
        catalog_id=catalog.id,
    )
    snapshot = snapshot_catalog(catalog=draft_catalog, items=items, plans=plans)
    draft_version = await create_catalog_version(
        session,
        catalog=catalog,
        snapshot=snapshot,
        created_by=actor_user_id,
        status="draft",
    )
    catalog.status = "draft"
    return {
        "created_count": created_count,
        "updated_count": updated_count,
        "draft_version_id": str(draft_version.id),
        "status": "draft_saved",
    }


def build_export_content(
    *,
    catalog: Catalog,
    snapshot: Mapping[str, Any],
    file_type: str,
    filters: Mapping[str, Any],
) -> tuple[bytes, str]:
    export_rows = _filter_snapshot_items(snapshot, filters)
    timestamp = datetime.now(UTC).strftime("%Y%m%d%H%M%S")
    base_name = f"{_normalize_slug(catalog.name) or 'catalog'}-{timestamp}"
    if file_type == "json":
        body = json.dumps(
            {
                "catalog": dict(snapshot.get("catalog") or {}),
                "items": export_rows,
                "filters": dict(filters),
            },
            ensure_ascii=False,
            indent=2,
        ).encode("utf-8")
        return body, f"{base_name}.json"
    tabular_rows = _flatten_export_rows(export_rows)
    if file_type == "csv":
        buffer = StringIO()
        writer = csv.DictWriter(buffer, fieldnames=sorted({key for row in tabular_rows for key in row}))
        writer.writeheader()
        for row in tabular_rows:
            writer.writerow(row)
        return buffer.getvalue().encode("utf-8"), f"{base_name}.csv"
    if file_type == "xlsx":
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Catalog"
        headers = sorted({key for row in tabular_rows for key in row})
        sheet.append(headers)
        for row in tabular_rows:
            sheet.append([row.get(header) for header in headers])
        output = BytesIO()
        workbook.save(output)
        return output.getvalue(), f"{base_name}.xlsx"
    raise ValueError("export file_type must be csv, xlsx or json")


def _filter_snapshot_items(snapshot: Mapping[str, Any], filters: Mapping[str, Any]) -> list[dict[str, Any]]:
    items = snapshot.get("items") if isinstance(snapshot.get("items"), list) else []
    selected: list[dict[str, Any]] = []
    active_only = bool(filters.get("active_only"))
    without_price = bool(filters.get("without_price"))
    with_plans = bool(filters.get("with_plans"))
    branch_id = _none_text(filters.get("branch_id"))
    category = _none_text(filters.get("category"))
    for raw_item in items:
        if not isinstance(raw_item, Mapping):
            continue
        item = dict(raw_item)
        if active_only and _normalize_slug(str(item.get("status") or "draft")) != "active":
            continue
        if without_price:
            has_price = _decimal_value(item.get("base_price")) is not None or _decimal_value(item.get("list_price")) is not None
            if has_price:
                continue
        if with_plans and not item.get("plans"):
            continue
        if branch_id and _none_text(item.get("branch_id")) != branch_id:
            continue
        if category and _none_text(item.get("category")) != category:
            continue
        selected.append(item)
    return selected


def _flatten_export_rows(items: Sequence[Mapping[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for item in items:
        base = {
            "sku": item.get("sku"),
            "name": item.get("name"),
            "type": item.get("type"),
            "category": item.get("category"),
            "base_price": item.get("base_price"),
            "list_price": item.get("list_price"),
            "stock_status": item.get("stock_status"),
            "stock_quantity": item.get("stock_quantity"),
            "branch_id": item.get("branch_id"),
            "status": item.get("status"),
            "tags_json": json.dumps(item.get("tags_json") or [], ensure_ascii=False),
            "attributes_json": json.dumps(item.get("attributes_json") or {}, ensure_ascii=False),
            "ai_rules_json": json.dumps(item.get("ai_rules_json") or {}, ensure_ascii=False),
        }
        plans = item.get("plans") if isinstance(item.get("plans"), list) else []
        if not plans:
            rows.append(base)
            continue
        for plan in plans:
            if not isinstance(plan, Mapping):
                continue
            rows.append(
                {
                    **base,
                    "plan_name": plan.get("plan_name"),
                    "plan_code": plan.get("plan_code"),
                    "plan_type": plan.get("plan_type"),
                    "down_payment_amount": plan.get("down_payment_amount"),
                    "down_payment_percent": plan.get("down_payment_percent"),
                    "installment_amount": plan.get("installment_amount"),
                    "installment_frequency": plan.get("installment_frequency"),
                    "installment_count": plan.get("installment_count"),
                    "term_months": plan.get("term_months"),
                    "eligibility_rules_json": json.dumps(
                        plan.get("eligibility_rules_json") or {},
                        ensure_ascii=False,
                    ),
                }
            )
    return rows


def _transform_import_row(
    raw_row: Mapping[str, Any],
    column_mapping: Mapping[str, str],
    *,
    default_status: str,
) -> tuple[ImportedItem, list[str]]:
    item_payload: dict[str, Any] = {
        "attributes_json": {},
        "ai_rules_json": {},
        "tags_json": [],
        "status": default_status,
        "stock_status": "unknown",
    }
    plans: dict[str, dict[str, Any]] = defaultdict(dict)
    errors: list[str] = []
    for column_name, target_path in column_mapping.items():
        if column_name not in raw_row:
            continue
        raw_value = raw_row.get(column_name)
        if raw_value in (None, ""):
            continue
        if target_path.startswith("plans."):
            _, plan_code, field_name = target_path.split(".", 2)
            plans[plan_code]["plan_code"] = plan_code
            plans[plan_code][field_name] = raw_value
            plans[plan_code].setdefault("plan_name", plan_code)
            inferred_frequency = _infer_installment_frequency(column_name)
            if inferred_frequency and "installment_frequency" not in plans[plan_code]:
                plans[plan_code]["installment_frequency"] = inferred_frequency
            continue
        if target_path.startswith("attributes."):
            item_payload["attributes_json"][target_path.split(".", 1)[1]] = _jsonish_value(raw_value)
            continue
        if target_path.startswith("ai_rules."):
            item_payload["ai_rules_json"][target_path.split(".", 1)[1]] = _jsonish_value(raw_value)
            continue
        if target_path == "tags_json":
            item_payload["tags_json"] = _listish_value(raw_value)
            continue
        if target_path in TOP_LEVEL_ITEM_FIELDS:
            item_payload[target_path] = raw_value
        else:
            item_payload["attributes_json"][target_path] = _jsonish_value(raw_value)
    _preserve_unmapped_import_attributes(raw_row, column_mapping, item_payload)
    _apply_nested_catalog_fallbacks(raw_row, item_payload, plans)
    item = ImportedItem(
        sku=str(item_payload.get("sku") or "").strip(),
        name=str(item_payload.get("name") or "").strip(),
        type=_none_text(item_payload.get("type")),
        category=_none_text(item_payload.get("category")),
        stock_status=_normalize_slug(str(item_payload.get("stock_status") or "unknown")) or "unknown",
        stock_quantity=_int_value(item_payload.get("stock_quantity")),
        branch_id=_none_text(item_payload.get("branch_id")),
        status=_normalize_slug(str(item_payload.get("status") or default_status)) or default_status,
        attributes_json=validate_json_dict(item_payload.get("attributes_json"), field_name="attributes_json"),
        ai_rules_json=validate_json_dict(item_payload.get("ai_rules_json"), field_name="ai_rules_json"),
        tags_json=[str(tag) for tag in validate_json_list(item_payload.get("tags_json"), field_name="tags_json")],
    )
    try:
        item.base_price = validate_price(item_payload.get("base_price"), field_name="base_price")
        item.list_price = validate_price(item_payload.get("list_price"), field_name="list_price")
        item.stock_status = validate_status(
            item.stock_status,
            allowed=STOCK_STATUSES,
            field_name="stock_status",
        )
        item.status = validate_status(item.status, allowed=CATALOG_STATUSES, field_name="status")
    except ValueError as exc:
        errors.append(str(exc))
    if not item.sku:
        errors.append("sku is required")
    if not item.name:
        errors.append("name is required")
    for plan_code, raw_plan in plans.items():
        try:
            plan = ImportedPlan(
                plan_name=str(raw_plan.get("plan_name") or plan_code).strip(),
                plan_code=str(raw_plan.get("plan_code") or plan_code).strip(),
                plan_type=_none_text(raw_plan.get("plan_type")),
                down_payment_amount=validate_price(
                    raw_plan.get("down_payment_amount"),
                    field_name=f"plan {plan_code} down_payment_amount",
                ),
                down_payment_percent=validate_price(
                    raw_plan.get("down_payment_percent"),
                    field_name=f"plan {plan_code} down_payment_percent",
                ),
                installment_amount=validate_price(
                    raw_plan.get("installment_amount"),
                    field_name=f"plan {plan_code} installment_amount",
                ),
                installment_frequency=_none_text(raw_plan.get("installment_frequency")),
                installment_count=validate_positive_int(
                    raw_plan.get("installment_count"),
                    field_name=f"plan {plan_code} installment_count",
                ),
                term_months=_int_value(raw_plan.get("term_months")),
                eligibility_rules_json=validate_json_dict(
                    raw_plan.get("eligibility_rules_json"),
                    field_name=f"plan {plan_code} eligibility_rules_json",
                ),
                status=validate_status(
                    str(raw_plan.get("status") or item.status),
                    allowed=CATALOG_STATUSES,
                    field_name=f"plan {plan_code} status",
                ),
            )
            if plan.installment_frequency and plan.installment_frequency not in PLAN_FREQUENCIES:
                errors.append(f"plan {plan_code} installment_frequency is invalid")
            item.plans.append(plan)
        except ValueError as exc:
            errors.append(str(exc))
    return item, errors


def _preserve_unmapped_import_attributes(
    raw_row: Mapping[str, Any],
    column_mapping: Mapping[str, str],
    item_payload: dict[str, Any],
) -> None:
    attrs = item_payload.setdefault("attributes_json", {})
    for column_name, raw_value in raw_row.items():
        if column_name in column_mapping:
            continue
        if raw_value in (None, ""):
            continue
        attrs.setdefault(str(column_name), _jsonish_value(raw_value))


def _apply_nested_catalog_fallbacks(
    raw_row: Mapping[str, Any],
    item_payload: dict[str, Any],
    plans: dict[str, dict[str, Any]],
) -> None:
    attrs = item_payload.setdefault("attributes_json", {})
    ai_rules = item_payload.setdefault("ai_rules_json", {})

    cash_price = item_payload.get("base_price")
    if cash_price in (None, ""):
        cash_price = raw_row.get("precio_contado_mxn")
    precios_mxn = raw_row.get("precios_mxn")
    if cash_price in (None, "") and isinstance(precios_mxn, Mapping):
        cash_price = precios_mxn.get("contado")
    if cash_price not in (None, ""):
        item_payload["base_price"] = cash_price
        attrs.setdefault("precio_contado_mxn", _jsonish_value(cash_price))
        attrs.setdefault("cash_price_mxn", _jsonish_value(cash_price))

    list_price = item_payload.get("list_price")
    if list_price in (None, ""):
        list_price = raw_row.get("precio_lista_mxn")
    if list_price in (None, "") and isinstance(precios_mxn, Mapping):
        list_price = precios_mxn.get("lista")
    if list_price not in (None, ""):
        item_payload["list_price"] = list_price
        attrs.setdefault("precio_lista_mxn", _jsonish_value(list_price))
        attrs.setdefault("list_price_mxn", _jsonish_value(list_price))

    for key in (
        "ficha_tecnica",
        "precios_mxn",
        "planes_credito",
        "planes_credito_normalizados",
        "modelo_moto",
        "busqueda_texto",
        "texto_retrieval",
        "tipo_registro",
        "campo_cliente_principal",
    ):
        raw_value = raw_row.get(key)
        if raw_value not in (None, ""):
            attrs.setdefault(key, _jsonish_value(raw_value))

    ficha = raw_row.get("ficha_tecnica")
    if isinstance(ficha, Mapping):
        attrs.setdefault("product_details", deepcopy(dict(ficha)))

    alias = _listish_value(raw_row.get("alias"))
    if alias:
        attrs.setdefault("alias", alias)
        attrs.setdefault("aliases", alias)

    alias_normalizados = _listish_value(raw_row.get("alias_normalizados"))
    if alias_normalizados:
        attrs.setdefault("alias_normalizados", alias_normalizados)
        attrs.setdefault("aliases", alias_normalizados if not alias else alias)

    if not item_payload.get("tags_json"):
        tags = _listish_value(raw_row.get("tags_uso"))
        if tags:
            item_payload["tags_json"] = tags
            attrs.setdefault("tags_uso", tags)

    if not plans:
        nested_plans = _coerce_nested_plans(
            raw_row.get("planes_credito_normalizados") or raw_row.get("planes_credito")
        )
        plans.update(nested_plans)
    if plans:
        ai_rules.setdefault("can_quote", True)
        ai_rules.setdefault("requires_plan", True)


def _coerce_nested_plans(raw_value: Any) -> dict[str, dict[str, Any]]:
    if not isinstance(raw_value, Mapping):
        return {}
    plans: dict[str, dict[str, Any]] = {}
    for raw_key, raw_plan in raw_value.items():
        if not isinstance(raw_plan, Mapping):
            continue
        plan_code = _normalized_plan_code(raw_key, raw_plan)
        eligibility_rules: dict[str, Any] = {}
        for field_name in ("plazo_texto", "respuesta_corta"):
            value = raw_plan.get(field_name)
            if value not in (None, ""):
                eligibility_rules[field_name] = value
        plans[plan_code] = {
            "plan_name": _none_text(raw_plan.get("plan_name"))
            or _none_text(raw_plan.get("plan"))
            or plan_code,
            "plan_code": plan_code,
            "down_payment_amount": raw_plan.get("enganche_mxn")
            or raw_plan.get("down_payment_amount")
            or raw_plan.get("down_payment_mxn"),
            "down_payment_percent": raw_plan.get("porcentaje_enganche")
            or raw_plan.get("down_payment_percent"),
            "installment_amount": raw_plan.get("pago_quincenal_mxn")
            or raw_plan.get("pago_mensual_mxn")
            or raw_plan.get("installment_amount")
            or raw_plan.get("installment_mxn"),
            "installment_frequency": _nested_plan_frequency(raw_plan),
            "installment_count": raw_plan.get("numero_quincenas")
            or raw_plan.get("numero_pagos")
            or raw_plan.get("installment_count")
            or raw_plan.get("term_count"),
            "term_months": raw_plan.get("term_months") or raw_plan.get("plazo_meses"),
            "eligibility_rules_json": eligibility_rules,
            "status": "active",
        }
    return plans


def _normalized_plan_code(raw_key: Any, raw_plan: Mapping[str, Any]) -> str:
    percentage = _int_value(raw_plan.get("porcentaje_enganche"))
    if percentage is not None:
        return f"{percentage}%"
    key_text = str(raw_key).strip()
    match = re.fullmatch(r"plan[_\s-]*(\d+)", key_text, flags=re.IGNORECASE)
    if match:
        return f"{match.group(1)}%"
    return key_text


def _nested_plan_frequency(raw_plan: Mapping[str, Any]) -> str | None:
    explicit = _none_text(raw_plan.get("installment_frequency")) or _none_text(raw_plan.get("frequency"))
    if explicit:
        normalized = _normalize_slug(explicit)
        if normalized in PLAN_FREQUENCIES:
            return normalized
    if raw_plan.get("pago_quincenal_mxn") not in (None, "") or raw_plan.get("numero_quincenas") not in (None, ""):
        return "quincenal"
    if raw_plan.get("pago_mensual_mxn") not in (None, ""):
        return "monthly"
    return None


def _extract_json_rows(parsed: Any) -> list[dict[str, Any]]:
    if isinstance(parsed, list):
        return [dict(item) for item in parsed if isinstance(item, Mapping)]
    if isinstance(parsed, Mapping):
        for key in ("rows", "items", "data", "catalog", "modelos", "products"):
            value = parsed.get(key)
            if isinstance(value, list):
                return [dict(item) for item in value if isinstance(item, Mapping)]
    return []


def _collect_columns(rows: Sequence[Mapping[str, Any]]) -> list[str]:
    seen: list[str] = []
    for row in rows:
        for key in row.keys():
            if key not in seen:
                seen.append(str(key))
    return seen


def _normalize_slug(value: str) -> str:
    normalized = NAME_RE.sub("_", (value or "").strip().lower())
    normalized = normalized.replace("-", "_")
    return normalized.strip("_")


def _normalize_import_key(value: str) -> str:
    normalized = _normalize_slug(value).replace(".", "_")
    normalized = IMPORT_KEY_SANITIZE_RE.sub("_", normalized)
    return normalized.strip("_")


def _detect_plan_column_target(column_name: str) -> str | None:
    normalized = _normalize_import_key(column_name)
    tokens = [token for token in normalized.split("_") if token]
    if len(tokens) < 2:
        return None
    plan_code = tokens[-1]
    if not any(character.isdigit() for character in plan_code):
        return None
    prefix = "_".join(tokens[:-1])
    plan_field_map = {
        "plan": "plan_name",
        "plan_name": "plan_name",
        "codigo_plan": "plan_code",
        "plan_code": "plan_code",
        "tipo_plan": "plan_type",
        "plan_type": "plan_type",
        "advance_payment": "down_payment_amount",
        "initial_payment": "down_payment_amount",
        "down_payment": "down_payment_amount",
        "down_payment_amount": "down_payment_amount",
        "down_payment_percent": "down_payment_percent",
        "installment": "installment_amount",
        "periodic_payment": "installment_amount",
        "biweekly_payment": "installment_amount",
        "monthly_payment": "installment_amount",
        "weekly_payment": "installment_amount",
        "daily_payment": "installment_amount",
        "installment_amount": "installment_amount",
        "frequency": "installment_frequency",
        "cadence": "installment_frequency",
        "installment_frequency": "installment_frequency",
        "payments": "installment_count",
        "term_count": "installment_count",
        "installment_count": "installment_count",
        "term_months": "term_months",
        "months": "term_months",
    }
    field_name = plan_field_map.get(prefix)
    if field_name is None:
        return None
    return f"plans.{plan_code}.{field_name}"


def _decimal_value(value: Any) -> Decimal | None:
    if value in (None, ""):
        return None
    if isinstance(value, Decimal):
        return value
    try:
        return Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, ValueError, TypeError):
        return None


def _decimal_text(value: Decimal | None) -> str | None:
    if value is None:
        return None
    text = format(value, "f")
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return text


def _decimal_number(value: Any) -> int | float | None:
    dec = _decimal_value(value)
    if dec is None:
        return None
    if dec == dec.to_integral():
        return int(dec)
    return float(dec)


def _int_value(value: Any) -> int | None:
    if value in (None, ""):
        return None
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return None


def _none_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _listish_value(value: Any) -> list[Any]:
    if value is None:
        return []
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return []
        if stripped.startswith("["):
            try:
                parsed = json.loads(stripped)
            except json.JSONDecodeError:
                parsed = None
            if isinstance(parsed, list):
                return parsed
        return [part.strip() for part in stripped.split(",") if part.strip()]
    if isinstance(value, Sequence) and not isinstance(value, (str, bytes, bytearray)):
        return list(value)
    return [value]


def _jsonish_value(value: Any) -> Any:
    if isinstance(value, str):
        stripped = value.strip()
        if stripped[:1] in {"{", "["}:
            try:
                return json.loads(stripped)
            except json.JSONDecodeError:
                return value
    return value


def _infer_installment_frequency(column_name: str) -> str | None:
    normalized = _normalize_import_key(column_name)
    if "quincenal" in normalized or "biweekly" in normalized:
        return "quincenal"
    if "mensual" in normalized or "monthly" in normalized:
        return "monthly"
    if "semanal" in normalized or "weekly" in normalized:
        return "weekly"
    if "diario" in normalized or "daily" in normalized:
        return "daily"
    return None
